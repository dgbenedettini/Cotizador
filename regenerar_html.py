import requests
import json
import pandas as pd
import openpyxl
import os
from io import BytesIO

IVA = 1.1

VENTAS_ID     = "1jXzje3Rg3reYCPfvZDk06kvBnPBO6WDq"
COMPRAS_ID    = "1B-JKT1VgnnYnEMarUzfKkjdkXTzzLOK1"
COTIZADOR_ID  = "1sBGkGzZuuBtkMuzo0h1t6IbE7EcSil1X"

def download_from_gdrive(file_id):
    """Descarga el archivo desde Google Drive usando export de Google Sheets (siempre actualizado)"""
    session = requests.Session()
    
    # Usar export de Google Sheets - siempre devuelve la version mas reciente
    url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
    r = session.get(url, stream=True, timeout=60)
    r.raise_for_status()
    
    content = r.content
    print(f"  Descargado: {len(content):,} bytes, tipo: {r.headers.get('Content-Type','?')}")
    
    # Verificar que es un archivo Excel (empieza con PK = ZIP)
    if content[:2] != b'PK':
        raise ValueError(f"Archivo descargado no es Excel. Primeros bytes: {content[:20]}")
    
    return BytesIO(content)

def load_ventas(file_obj):
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    print(f"  Hojas disponibles: {wb.sheetnames}")
    ws = wb['VENTAS']
    rows, header = [], None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 19:
            header = list(row[:30])
        elif i > 19:
            if any(x is not None for x in row):
                rows.append(row[:30])
    df = pd.DataFrame(rows, columns=header)
    df = df[df['CAMPAÑA'] == '26-27'].copy()
    for col in ['PRECIO','PRECIO S/IVA','VOLUMEN','TOTAL C/IVA']:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    return df, wb

def load_compras(file_obj):
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    print(f"  Hojas disponibles: {wb.sheetnames}")
    ws = wb['COMPRAS']
    rows, header = [], None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 15:
            header = list(row[:20])
        elif i > 15:
            if any(x is not None for x in row):
                rows.append(row[:20])
    df = pd.DataFrame(rows, columns=header)
    df['VOLUMEN'] = pd.to_numeric(df['VOLUMEN'], errors='coerce')
    df['COSTO']   = pd.to_numeric(df['COSTO'],   errors='coerce')
    return df

def load_descripciones(file_obj):
    wb = openpyxl.load_workbook(file_obj, read_only=True, data_only=True)
    ws = wb['COTIZADOR']
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 7 and any(x is not None for x in row):
            rows.append(row[:8])
    df = pd.DataFrame(rows, columns=['_','COMPANIA','PRODUCTO','DESCRIPCION','COSTO','PRECIO_VENTA','MB','VOLUMEN'])
    df = df[df['PRODUCTO'].notna() & (df['PRODUCTO'] != 'Producto')].copy()
    descripciones = {}
    for _, row in df.iterrows():
        prod = str(row['PRODUCTO']).strip().upper()
        desc = str(row['DESCRIPCION']).strip() if pd.notna(row['DESCRIPCION']) and row['DESCRIPCION'] != 0 else ''
        if prod and desc:
            descripciones[prod] = desc
    print(f'  Descripciones cargadas: {len(descripciones)}')
    return descripciones

def calc_costo_ponderado(df_c):
    df = df_c[df_c['COSTO'].notna() & df_c['VOLUMEN'].notna()].copy()
    df['TOTAL_COSTO'] = df['VOLUMEN'] * df['COSTO']
    cp = df.groupby('PRODUCTO').apply(
        lambda x: round(x['TOTAL_COSTO'].sum() / x['VOLUMEN'].sum(), 4)
        if x['VOLUMEN'].sum() != 0 else None
    ).reset_index()
    cp.columns = ['PRODUCTO', 'COSTO_POND']
    return cp[cp['COSTO_POND'].notna()]

def calc_stock(wb_v):
    ws_s = wb_v['STOCK']
    rows = []
    for i, row in enumerate(ws_s.iter_rows(values_only=True)):
        if i > 6 and any(x is not None for x in row):
            rows.append(row[:12])
    cols = [None,'CAMPAÑA','FECHA','MOVIMIENTO','PRODUCTO','LAB','VOLUMEN','IN','OUT','CLIENTE','OBS','STATUS']
    df = pd.DataFrame(rows, columns=cols)
    df['VOLUMEN'] = pd.to_numeric(df['VOLUMEN'], errors='coerce')
    stock = df.groupby('PRODUCTO')['VOLUMEN'].sum().reset_index()
    stock.columns = ['PRODUCTO','STOCK']
    return stock

def build_app_data(df_v, df_c, wb_v, descripciones):
    cp    = calc_costo_ponderado(df_c)
    stock = calc_stock(wb_v)
    catalog = cp.merge(stock, on='PRODUCTO', how='left')
    catalog['STOCK'] = catalog['STOCK'].fillna(0).round(1)
    lab_map = df_c[df_c['COSTO'].notna()].groupby('PRODUCTO')['COMPAÑÍA'].first().reset_index()
    catalog = catalog.merge(lab_map, on='PRODUCTO', how='left')

    catalog_json = []
    for _, r in catalog.iterrows():
        nombre = str(r['PRODUCTO']).strip().upper()
        catalog_json.append({
            'producto': r['PRODUCTO'],
            'costo':    round(float(r['COSTO_POND']), 4),
            'stock':    float(r['STOCK']),
            'lab':      str(r['COMPAÑÍA']) if pd.notna(r['COMPAÑÍA']) else '',
            'desc':     descripciones.get(nombre, '')
        })

    df_hist = df_v.merge(cp, on='PRODUCTO', how='left')

    def calc_mb(row):
        p, c = row['PRECIO'], row['COSTO_POND']
        if pd.isna(p) or pd.isna(c) or p == 0:
            return None
        mb = (p/IVA - c/IVA) / p * 100
        return round(mb, 1) if (mb == mb) and abs(mb) != float('inf') else None

    df_hist['MB_PCT'] = df_hist.apply(calc_mb, axis=1)

    ventas_json = {}
    for _, row in df_hist.iterrows():
        cliente = str(row['CLIENTE']) if pd.notna(row['CLIENTE']) else None
        if not cliente:
            continue
        if cliente not in ventas_json:
            ventas_json[cliente] = []
        fecha_str = ''
        if pd.notna(row['FECHA']):
            try:
                fecha_str = row['FECHA'].strftime('%d/%m/%Y') if hasattr(row['FECHA'],'strftime') else str(row['FECHA'])
            except:
                fecha_str = str(row['FECHA'])
        ventas_json[cliente].append({
            'fecha':     fecha_str,
            'producto':  str(row['PRODUCTO'])      if pd.notna(row['PRODUCTO'])      else '',
            'lab':       str(row['COMPAÑÍA'])      if pd.notna(row['COMPAÑÍA'])      else '',
            'volumen':   float(row['VOLUMEN'])     if pd.notna(row['VOLUMEN'])       else None,
            'precio':    float(row['PRECIO'])      if pd.notna(row['PRECIO'])        else None,
            'costo':     float(row['COSTO_POND'])  if pd.notna(row['COSTO_POND'])    else None,
            'mb_pct':    row['MB_PCT'],
            'condicion': str(row['CONDICION'])     if pd.notna(row['CONDICION'])     else '',
            'total':     float(row['TOTAL C/IVA']) if pd.notna(row['TOTAL C/IVA'])   else None
        })

    return {
        'catalogo': catalog_json,
        'clientes': sorted(ventas_json.keys()),
        'ventas':   ventas_json
    }

def update_html(app_data):
    with open('index.html', encoding='utf-8') as f:
        html = f.read()
    raw = json.dumps(app_data, ensure_ascii=False, separators=(',',':'))
    s = html.find('const APP_DATA=') + len('const APP_DATA=')
    e = html.find(';\n', s)
    html = html[:s] + raw + html[e:]
    with open('index.html', 'w', encoding='utf-8') as f:
        f.write(html)
    print(f'index.html actualizado — {len(html):,} chars')

def main():
    print('Descargando VENTAS desde Google Drive...')
    ventas_file  = download_from_gdrive(VENTAS_ID)
    print('Descargando COMPRAS desde Google Drive...')
    compras_file = download_from_gdrive(COMPRAS_ID)

    print('Cargando VENTAS...')
    df_v, wb_v = load_ventas(ventas_file)
    print('Cargando COMPRAS...')
    df_c = load_compras(compras_file)

    print('Descargando COTIZADOR desde Google Drive...')
    cotizador_file = download_from_gdrive(COTIZADOR_ID)
    print('Cargando descripciones...')
    descripciones = load_descripciones(cotizador_file)
    # Guardar tambien como JSON para respaldo
    with open('descripciones.json', 'w', encoding='utf-8') as f:
        json.dump(descripciones, f, ensure_ascii=False)

    print('Calculando datos...')
    app_data = build_app_data(df_v, df_c, wb_v, descripciones)
    update_html(app_data)
    print(f'Listo. Productos: {len(app_data["catalogo"])}, Clientes: {len(app_data["clientes"])}')

if __name__ == '__main__':
    main()
